# ExcelPDFPortable.py
# 스크린샷과 동일 레이아웃 + 모든 기능(시트복사/이어붙이기/통합엑셀/각종 PDF 출력) 포함
# Windows + Microsoft Excel 권장(서식 보존 복사 & PDF 내보내기용)
import os, sys, tempfile, shutil
from collections import defaultdict
from PyQt5 import QtCore, QtGui, QtWidgets

EXCEL_EXTS = {".xls", ".xlsx", ".xlsm", ".xlsb"}

def is_excel(path: str) -> bool:
    return os.path.splitext(path)[1].lower() in EXCEL_EXTS

# ------------------------ Excel COM (서식 보존 & PDF) ------------------------
class ExcelCom:
    def __init__(self):
        self.excel = None

    def _ensure(self):
        if self.excel is not None:
            return True
        try:
            import win32com.client as win32
            self.win32 = win32
            self.excel = win32.gencache.EnsureDispatch("Excel.Application")
            self.excel.Visible = False
            self.excel.DisplayAlerts = False
            return True
        except Exception as e:
            self.excel = None
            return False

    def close(self):
        try:
            if self.excel:
                self.excel.DisplayAlerts = False
                self.excel.Quit()
        except Exception:
            pass
        self.excel = None

    # 열린/닫힌 파일 안전하게 열기
    def open_wb(self, path):
        wb = self.excel.Workbooks.Open(os.path.abspath(path))
        return wb

    def new_wb(self):
        return self.excel.Workbooks.Add()

    def save_wb_as(self, wb, path):
        # 51 = xlOpenXMLWorkbook(.xlsx), 56 = xls
        ext = os.path.splitext(path)[1].lower()
        if ext == ".xlsx":
            fmt = 51
        elif ext == ".xlsm":
            fmt = 52
        elif ext == ".xls":
            fmt = 56
        else:
            fmt = 51
        wb.SaveAs(os.path.abspath(path), FileFormat=fmt)

    def export_pdf(self, wb, out_pdf_path, sheet_names=None, one_pdf=True):
        """
        sheet_names가 None이면 통으로, 아니면 지정 시트만.
        one_pdf=True면 한 개 PDF, False면 시트별로 각각(호출부에서 반복 사용 권장).
        """
        out_pdf_path = os.path.abspath(out_pdf_path)
        # Type=0 : xlTypePDF
        xlTypePDF = 0
        if sheet_names:
            shts = [wb.Worksheets(s) for s in sheet_names]
            # 여러 시트 선택 후 Export
            wb.Worksheets(1).Select(False)
            shts[0].Select(False)
            for s in shts[1:]:
                s.Select(True)
        wb.ActiveSheet.ExportAsFixedFormat(Type=xlTypePDF,
                                           Filename=out_pdf_path,
                                           Quality=0,  # Standard
                                           IncludeDocProperties=True,
                                           IgnorePrintAreas=False,
                                           OpenAfterPublish=False)

    def copy_sheet_to(self, src_path, sheet_name, dst_wb):
        """서식 포함 시트 복사: src 열고 해당 시트 Copy After=dst 마지막 시트"""
        src_wb = None
        try:
            src_wb = self.open_wb(src_path)
            sheet = src_wb.Worksheets(sheet_name)
            # 대상 워크북이 비어 있으면 그냥 Copy로 생성
            if dst_wb.Worksheets.Count == 1 and dst_wb.Worksheets(1).UsedRange.Count == 1 and not dst_wb.Worksheets(1).Name:
                sheet.Copy(Before=dst_wb.Worksheets(1))
                # 기본 첫 빈 시트 제거
                try:
                    dst_wb.Worksheets(2).Delete()
                except Exception:
                    pass
            else:
                sheet.Copy(After=dst_wb.Worksheets(dst_wb.Worksheets.Count))
        finally:
            if src_wb:
                src_wb.Close(SaveChanges=False)

# ------------------------ 파일 리스트 위젯 ------------------------
class FileList(QtWidgets.QListWidget):
    filesChanged = QtCore.pyqtSignal()
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.setAlternatingRowColors(True)
        self.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self._menu)

    def dragEnterEvent(self, e: QtGui.QDragEnterEvent):
        if e.mimeData().hasUrls(): e.acceptProposedAction()
        else: e.ignore()

    def dropEvent(self, e: QtGui.QDropEvent):
        added = 0
        for u in e.mimeData().urls():
            p = u.toLocalFile()
            if os.path.isdir(p):
                for r, _, fs in os.walk(p):
                    for f in fs:
                        fp = os.path.join(r, f)
                        if is_excel(fp): added += self._add_unique(fp)
            elif is_excel(p):
                added += self._add_unique(p)
        if added:
            self.sortItems()
            self.filesChanged.emit()

    def keyPressEvent(self, e: QtGui.QKeyEvent):
        if e.key() in (QtCore.Qt.Key_Delete, QtCore.Qt.Key_Backspace):
            self.remove_selected()
        else:
            super().keyPressEvent(e)

    def _menu(self, pos):
        m = QtWidgets.QMenu(self)
        a1 = m.addAction("선택 항목 제거")
        a2 = m.addAction("전체 비우기")
        act = m.exec_(self.mapToGlobal(pos))
        if act == a1: self.remove_selected()
        elif act == a2:
            self.clear(); self.filesChanged.emit()

    def _add_unique(self, path: str) -> int:
        path = os.path.normpath(path)
        for i in range(self.count()):
            if self.item(i).data(QtCore.Qt.UserRole) == path: return 0
        it = QtWidgets.QListWidgetItem(os.path.basename(path))
        it.setToolTip(path)
        it.setData(QtCore.Qt.UserRole, path)
        it.setIcon(self.style().standardIcon(QtWidgets.QStyle.SP_FileIcon))
        self.addItem(it)
        return 1

    def remove_selected(self):
        for it in self.selectedItems():
            self.takeItem(self.row(it))
        self.filesChanged.emit()

    def paths(self):
        return [self.item(i).data(QtCore.Qt.UserRole) for i in range(self.count())]

# ------------------------ 메인 창 ------------------------
class Main(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("엑셀 병합 · PDF 변환 (포터블)")
        self.resize(1000, 700)
        try:
            import ctypes
            ctypes.windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            pass
        self.pdf_base_dir = ""

        # 1) 파일 추가
        t1 = QtWidgets.QLabel("1) 엑셀 파일 추가 — 드래그앤드롭을 가능")
        t1.setStyleSheet("font-weight:600;")
        self.fileList = FileList()
        self.file_status = QtWidgets.QLabel("0개 파일")
        self.fileList.filesChanged.connect(lambda: self.file_status.setText(f"{self.fileList.count()}개 파일"))

        b_add = QtWidgets.QPushButton("엑셀 파일 추가")
        b_add.clicked.connect(self.add_files)
        b_del = QtWidgets.QPushButton("선택 항목 제거")
        b_del.clicked.connect(self.fileList.remove_selected)
        b_clear = QtWidgets.QPushButton("전체 비우기")
        b_clear.clicked.connect(lambda: (self.fileList.clear(), self.fileList.filesChanged.emit()))
        b_load = QtWidgets.QPushButton("시트 목록 불러오기")
        b_load.clicked.connect(self.load_sheets)

        row1 = QtWidgets.QHBoxLayout()
        for b in (b_add, b_del, b_clear, b_load): row1.addWidget(b)
        row1.addStretch(); row1.addWidget(self.file_status)

        # 2) 시트 선택 및 순서
        t2 = QtWidgets.QLabel("2) 시트 선택 및 순서 지정")
        t2.setStyleSheet("font-weight:600;")
        self.sheetLeft = QtWidgets.QListWidget()
        self.sheetRight = QtWidgets.QListWidget()
        for lw in (self.sheetLeft, self.sheetRight):
            lw.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
            lw.setAlternatingRowColors(True)

        cap = QtWidgets.QHBoxLayout()
        cap.addWidget(QtWidgets.QLabel("사용 가능(왼쪽) → 선택(오른쪽)")); cap.addStretch()

        to_right = QtWidgets.QPushButton("추가 ▶"); to_right.clicked.connect(self.move_to_right)
        to_left  = QtWidgets.QPushButton("◀ 제거"); to_left.clicked.connect(self.move_to_left)
        up_btn   = QtWidgets.QPushButton("위로 ↑"); up_btn.clicked.connect(lambda: self.move_up_down(-1))
        dn_btn   = QtWidgets.QPushButton("아래로 ↓"); dn_btn.clicked.connect(lambda: self.move_up_down(+1))
        clr_sel  = QtWidgets.QPushButton("모두 제거"); clr_sel.clicked.connect(self.sheetRight.clear)

        mid = QtWidgets.QVBoxLayout()
        for b in (to_right, to_left, up_btn, dn_btn, clr_sel): mid.addWidget(b)
        mid.addStretch()

        grids = QtWidgets.QHBoxLayout()
        grids.addWidget(self.sheetLeft, 1); grids.addLayout(mid); grids.addWidget(self.sheetRight, 1)

        # 3) 병합 모드
        t3 = QtWidgets.QLabel("3) 병합 모드:"); t3.setStyleSheet("font-weight:600;")
        self.rb_copy = QtWidgets.QRadioButton("시트 복사 (서식 보존, 기본)")
        self.rb_concat = QtWidgets.QRadioButton("데이터 이어붙이기 (한 시트)")
        self.rb_copy.setChecked(True)
        mergeRow = QtWidgets.QHBoxLayout()
        mergeRow.addWidget(t3); mergeRow.addWidget(self.rb_copy); mergeRow.addWidget(self.rb_concat); mergeRow.addStretch()

        # 4) PDF 폴더 & 출력방식
        self.lbl_pdf = QtWidgets.QLabel("4) PDF 기본 저장 폴더:  (미설정)")
        self.lbl_pdf.setStyleSheet("font-weight:600;")
        b_pdf_dir = QtWidgets.QPushButton("변경..."); b_pdf_dir.clicked.connect(self.pick_pdf_folder)
        pdfRow1 = QtWidgets.QHBoxLayout(); pdfRow1.addWidget(self.lbl_pdf); pdfRow1.addWidget(b_pdf_dir); pdfRow1.addStretch()

        self.rb_pdf_merged = QtWidgets.QRadioButton("통합 PDF 1개")
        self.rb_pdf_by_sheet = QtWidgets.QRadioButton("시트별 개별 PDF")
        self.rb_pdf_by_file = QtWidgets.QRadioButton("원본 파일별 PDF")
        self.rb_pdf_merged.setChecked(True)
        pdfRow2 = QtWidgets.QHBoxLayout()
        pdfRow2.addWidget(QtWidgets.QLabel("PDF 출력 방식:"))
        for rb in (self.rb_pdf_merged, self.rb_pdf_by_sheet, self.rb_pdf_by_file):
            pdfRow2.addWidget(rb)
        pdfRow2.addStretch()

        # 실행 버튼
        b_make_excel = QtWidgets.QPushButton("통합 엑셀 만들기")
        b_make_pdf   = QtWidgets.QPushButton("PDF 만들기")
        b_both       = QtWidgets.QPushButton("한 번에: 통합 엑셀 + PDF")
        b_make_excel.clicked.connect(self.action_make_excel)
        b_make_pdf.clicked.connect(self.action_make_pdf)
        b_both.clicked.connect(self.action_make_both)

        run = QtWidgets.QHBoxLayout()
        for b in (b_make_excel, b_make_pdf, b_both): run.addWidget(b)
        run.addStretch()

        # 전체 레이아웃
        lay = QtWidgets.QVBoxLayout(self)
        lay.addWidget(t1); lay.addWidget(self.fileList, 1); lay.addLayout(row1)
        lay.addWidget(t2); lay.addLayout(cap); lay.addLayout(grids)
        lay.addLayout(mergeRow)
        lay.addLayout(pdfRow1); lay.addLayout(pdfRow2)
        lay.addSpacing(6); lay.addLayout(run)

    # ---------- 유틸 ----------
    def info(self, m): QtWidgets.QMessageBox.information(self, "알림", m)
    def warn(self, m): QtWidgets.QMessageBox.warning(self, "안내", m)
    def err (self, m): QtWidgets.QMessageBox.critical(self, "오류", m)

    def add_files(self):
        files, _ = QtWidgets.QFileDialog.getOpenFileNames(
            self, "엑셀 파일 선택", "", "Excel Files (*.xlsx *.xls *.xlsm *.xlsb);;All Files (*.*)"
        )
        added = 0
        for f in files:
            if is_excel(f):
                added += self.fileList._add_unique(f)
        if added:
            self.fileList.sortItems()
            self.fileList.filesChanged.emit()

    def pick_pdf_folder(self):
        d = QtWidgets.QFileDialog.getExistingDirectory(self, "PDF 저장 폴더 선택", "")
        if d:
            self.pdf_base_dir = d
            self.lbl_pdf.setText(f"4) PDF 기본 저장 폴더:  {d}")

    # ---------- 시트 목록 ----------
    def load_sheets(self):
        self.sheetLeft.clear()
        paths = self.fileList.paths()
        if not paths:
            self.warn("먼저 엑셀 파일을 추가하세요.")
            return
        count = 0
        for p in paths:
            ext = os.path.splitext(p)[1].lower()
            try:
                if ext in (".xlsx", ".xlsm", ".xlsb"):
                    from openpyxl import load_workbook
                    wb = load_workbook(p, read_only=True, data_only=True)
                    for s in wb.sheetnames:
                        self._add_sheet_left(f"{os.path.basename(p)} | {s}", p, s); count += 1
                elif ext == ".xls":
                    import xlrd
                    wb = xlrd.open_workbook(p)
                    for s in wb.sheet_names():
                        self._add_sheet_left(f"{os.path.basename(p)} | {s}", p, s); count += 1
            except Exception as e:
                print("시트 로드 오류:", p, e)
        self.info(f"시트 {count}개를 불러왔습니다.")

    def _add_sheet_left(self, label, file_path, sheet_name):
        it = QtWidgets.QListWidgetItem(label)
        it.setData(QtCore.Qt.UserRole, (file_path, sheet_name))
        self.sheetLeft.addItem(it)

    def move_to_right(self):
        for it in self.sheetLeft.selectedItems():
            clone = QtWidgets.QListWidgetItem(it.text())
            clone.setData(QtCore.Qt.UserRole, it.data(QtCore.Qt.UserRole))
            self.sheetRight.addItem(clone)

    def move_to_left(self):
        for it in self.sheetRight.selectedItems():
            self.sheetRight.takeItem(self.sheetRight.row(it))

    def move_up_down(self, delta: int):
        lw = self.sheetRight
        rows = sorted([lw.row(i) for i in lw.selectedItems()])
        if not rows: return
        if delta < 0:
            for r in rows:
                if r == 0: continue
                it = lw.takeItem(r); lw.insertItem(r-1, it); lw.setItemSelected(it, True)
        else:
            for r in reversed(rows):
                if r == lw.count()-1: continue
                it = lw.takeItem(r); lw.insertItem(r+1, it); lw.setItemSelected(it, True)

    # ---------- 동작: 통합 엑셀 ----------
    def action_make_excel(self, also_return_path=False):
        items = [self.sheetRight.item(i) for i in range(self.sheetRight.count())]
        if not items:
            self.warn("오른쪽(선택) 목록에 시트를 추가하세요.")
            return None

        # 저장 경로
        save_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self, "통합 엑셀 저장 위치", "", "Excel Workbook (*.xlsx);;Excel Macro-Enabled (*.xlsm);;Excel 97-2003 (*.xls)"
        )
        if not save_path:
            return None

        if self.rb_copy.isChecked():
            # Excel COM으로 시트 복사(서식 보존)
            com = ExcelCom()
            if not com._ensure():
                self.err("Microsoft Excel이 필요합니다. Excel이 설치된 환경에서 실행해 주세요.")
                return None
            try:
                dst = com.new_wb()
                # 초기 빈 시트 이름 제거 위해 이름 비워둔 채 넣고 나중에 삭제 시도
                for idx, it in enumerate(items, start=1):
                    fp, sn = it.data(QtCore.Qt.UserRole)
                    com.copy_sheet_to(fp, sn, dst)
                    # 붙여넣은 시트의 이름 충돌 방지를 위해 파일명|시트명으로 변경 시도
                    try:
                        dst.Worksheets(dst.Worksheets.Count).Name = f"{os.path.splitext(os.path.basename(fp))[0]}_{sn}"[:31]
                    except Exception:
                        pass
                # 첫 번째 빈 시트가 남아있으면 제거
                try:
                    if dst.Worksheets(1).UsedRange.Count == 1 and not dst.Worksheets(1).Name:
                        dst.Worksheets(1).Delete()
                except Exception:
                    pass
                com.save_wb_as(dst, save_path)
                dst.Close(SaveChanges=False)
                self.info("통합 엑셀이 생성되었습니다.")
                return save_path if also_return_path else None
            except Exception as e:
                self.err(f"엑셀 시트 복사 중 오류: {e}")
                return None
            finally:
                com.close()

        else:
            # 이어붙이기: openpyxl 값 기반 병합
            try:
                from openpyxl import Workbook, load_workbook
                out_wb = Workbook()
                ws = out_wb.active
                ws.title = "MergedData"
                wrote_header = False
                row_cursor = 1
                for it in items:
                    fp, sn = it.data(QtCore.Qt.UserRole)
                    ext = os.path.splitext(fp)[1].lower()
                    if ext in (".xlsx", ".xlsm", ".xlsb"):
                        wb = load_workbook(fp, data_only=True)
                        sh = wb[sn]
                        data = list(sh.values)
                    else:
                        import xlrd
                        wb = xlrd.open_workbook(fp)
                        sh = wb.sheet_by_name(sn)
                        data = [sh.row_values(r) for r in range(sh.nrows)]
                    if not data: continue
                    # 헤더는 첫 파일의 첫 행만
                    start_idx = 0
                    if wrote_header:
                        start_idx = 1  # 첫 행을 헤더로 가정, 이후 파일은 헤더 스킵
                    else:
                        wrote_header = True
                    for r in data[start_idx:]:
                        for c_idx, v in enumerate(r, start=1):
                            ws.cell(row=row_cursor, column=c_idx, value=v)
                        row_cursor += 1
                out_wb.save(save_path)
                self.info("통합 엑셀이 생성되었습니다.")
                return save_path if also_return_path else None
            except Exception as e:
                self.err(f"데이터 이어붙이기 중 오류: {e}")
                return None

    # ---------- 동작: PDF ----------
    def action_make_pdf(self):
        items = [self.sheetRight.item(i) for i in range(self.sheetRight.count())]
        if not items:
            self.warn("오른쪽(선택) 목록에 시트를 추가하세요.")
            return

        if not self.pdf_base_dir:
            self.warn("PDF 저장 폴더를 먼저 설정하세요.")
            return

        if self.rb_copy.isChecked():
            # 우선 통합 엑셀 임시 생성 후 PDF 출력
            tmp_dir = tempfile.mkdtemp(prefix="excelpdf_")
            tmp_xlsx = os.path.join(tmp_dir, "merged.xlsx")
            try:
                # 통합 엑셀(서식보존) 임시 생성
                excel_path = self._make_excel_temp_for_pdf(tmp_xlsx, items)
                if not excel_path:
                    return
                # 출력 방식
                com = ExcelCom()
                if not com._ensure():
                    self.err("Microsoft Excel이 필요합니다. Excel이 설치된 환경에서 실행해 주세요.")
                    return
                wb = com.open_wb(excel_path)

                if self.rb_pdf_merged.isChecked():
                    out = os.path.join(self.pdf_base_dir, "merged.pdf")
                    # 모든 시트를 선택해서 1개 PDF
                    names = [wb.Worksheets(i).Name for i in range(1, wb.Worksheets.Count+1)]
                    wb.Worksheets(names[0]).Select(False)
                    for n in names[1:]:
                        wb.Worksheets(n).Select(True)
                    com.export_pdf(wb, out, sheet_names=None, one_pdf=True)
                    self.info(f"PDF 생성 완료: {out}")

                elif self.rb_pdf_by_sheet.isChecked():
                    # 시트별 각각
                    for i in range(1, wb.Worksheets.Count+1):
                        name = wb.Worksheets(i).Name
                        safe = "".join(ch for ch in name if ch not in '\\/:*?"<>|')
                        out = os.path.join(self.pdf_base_dir, f"{safe}.pdf")
                        wb.Worksheets(name).Select(False)
                        com.export_pdf(wb, out, sheet_names=[name], one_pdf=True)
                    self.info("시트별 PDF 생성 완료")

                else:  # 원본 파일별: 선택된 시트를 파일별로 묶어 PDF
                    file_groups = defaultdict(list)
                    for it in items:
                        fp, sn = it.data(QtCore.Qt.UserRole)
                        file_groups[fp].append(sn)
                    # 통합본에서 이름 규칙은 파일명_시트명으로 생성했으니, 매칭
                    for fp, sns in file_groups.items():
                        names_in_merged = []
                        for sn in sns:
                            nm = f\"{os.path.splitext(os.path.basename(fp))[0]}_{sn}\"[:31]
                            names_in_merged.append(nm)
                        # 안전한 출력명
                        base = os.path.splitext(os.path.basename(fp))[0]
                        out = os.path.join(self.pdf_base_dir, f\"{base}.pdf\")
                        wb.Worksheets(names_in_merged[0]).Select(False)
                        for n in names_in_merged[1:]:
                            wb.Worksheets(n).Select(True)
                        com.export_pdf(wb, out, sheet_names=None, one_pdf=True)
                    self.info("원본 파일별 PDF 생성 완료")

                wb.Close(SaveChanges=False)
                com.close()
            except Exception as e:
                self.err(f"PDF 생성 중 오류: {e}")
            finally:
                try:
                    shutil.rmtree(tmp_dir, ignore_errors=True)
                except Exception:
                    pass

        else:
            # 이어붙이기: 값 기반 한 시트 통합 후 PDF
            tmp_dir = tempfile.mkdtemp(prefix="excelpdf_")
            tmp_xlsx = os.path.join(tmp_dir, "merged.xlsx")
            try:
                excel_path = self._make_excel_concat_tmp(tmp_xlsx, items)
                if not excel_path:
                    return
                com = ExcelCom()
                if not com._ensure():
                    self.err("Microsoft Excel이 필요합니다. Excel이 설치된 환경에서 실행해 주세요.")
                    return
                wb = com.open_wb(excel_path)
                out = os.path.join(self.pdf_base_dir, "merged.pdf") if self.rb_pdf_merged.isChecked() else None

                if self.rb_pdf_merged.isChecked() or self.rb_pdf_by_file.isChecked():
                    if not out:
                        out = os.path.join(self.pdf_base_dir, "merged.pdf")
                    wb.Worksheets(1).Select(False)
                    com.export_pdf(wb, out, sheet_names=None, one_pdf=True)
                    self.info(f"PDF 생성 완료: {out}")
                else:  # 시트별(여긴 1개 시트뿐이므로 동일 결과)
                    wb.Worksheets(1).Select(False)
                    out = os.path.join(self.pdf_base_dir, "MergedData.pdf")
                    com.export_pdf(wb, out, sheet_names=None, one_pdf=True)
                    self.info(f"PDF 생성 완료: {out}")

                wb.Close(SaveChanges=False)
                com.close()
            except Exception as e:
                self.err(f"PDF 생성 중 오류: {e}")
            finally:
                try:
                    shutil.rmtree(tmp_dir, ignore_errors=True)
                except Exception:
                    pass

    # 임시 통합 엑셀(복사모드) 생성
    def _make_excel_temp_for_pdf(self, save_path, items):
        com = ExcelCom()
        if not com._ensure():
            self.err("Microsoft Excel이 필요합니다. Excel이 설치된 환경에서 실행해 주세요.")
            return None
        try:
            dst = com.new_wb()
            for it in items:
                fp, sn = it.data(QtCore.Qt.UserRole)
                com.copy_sheet_to(fp, sn, dst)
                try:
                    dst.Worksheets(dst.Worksheets.Count).Name = f\"{os.path.splitext(os.path.basename(fp))[0]}_{sn}\"[:31]
                except Exception:
                    pass
            com.save_wb_as(dst, save_path)
            dst.Close(SaveChanges=False)
            com.close()
            return save_path
        except Exception as e:
            com.close()
            self.err(f"통합본 작성 오류: {e}")
            return None

    # 임시 통합 엑셀(이어붙이기) 생성
    def _make_excel_concat_tmp(self, save_path, items):
        try:
            from openpyxl import Workbook, load_workbook
            import xlrd
        except Exception:
            pass
        try:
            from openpyxl import Workbook, load_workbook
            out_wb = Workbook(); ws = out_wb.active; ws.title = "MergedData"
            wrote_header = False; row_cursor = 1
            for it in items:
                fp, sn = it.data(QtCore.Qt.UserRole)
                ext = os.path.splitext(fp)[1].lower()
                if ext in (".xlsx", ".xlsm", ".xlsb"):
                    wb = load_workbook(fp, data_only=True)
                    sh = wb[sn]
                    data = list(sh.values)
                else:
                    import xlrd
                    wb = xlrd.open_workbook(fp)
                    sh = wb.sheet_by_name(sn)
                    data = [sh.row_values(r) for r in range(sh.nrows)]
                if not data: continue
                start_idx = 0
                if wrote_header:
                    start_idx = 1
                else:
                    wrote_header = True
                for r in data[start_idx:]:
                    for c_idx, v in enumerate(r, start=1):
                        ws.cell(row=row_cursor, column=c_idx, value=v)
                    row_cursor += 1
            out_wb.save(save_path)
            return save_path
        except Exception as e:
            self.err(f"이어붙이기 통합 작성 오류: {e}")
            return None

    # ---------- 동작: 한 번에 ----------
    def action_make_both(self):
        # 1) 통합 엑셀
        merged_path = self.action_make_excel(also_return_path=True)
        if not merged_path:
            return
        # 2) PDF (통합 파일 기반으로 출력)
        if not self.pdf_base_dir:
            self.warn("PDF 저장 폴더를 설정하고 다시 시도하세요.")
            return
        try:
            com = ExcelCom()
            if not com._ensure():
                self.err("Microsoft Excel이 필요합니다. Excel이 설치된 환경에서 실행해 주세요.")
                return
            wb = com.open_wb(merged_path)
            if self.rb_pdf_merged.isChecked():
                out = os.path.join(self.pdf_base_dir, "merged.pdf")
                names = [wb.Worksheets(i).Name for i in range(1, wb.Worksheets.Count+1)]
                wb.Worksheets(names[0]).Select(False)
                for n in names[1:]:
                    wb.Worksheets(n).Select(True)
                com.export_pdf(wb, out, sheet_names=None, one_pdf=True)
            elif self.rb_pdf_by_sheet.isChecked():
                for i in range(1, wb.Worksheets.Count+1):
                    name = wb.Worksheets(i).Name
                    safe = "".join(ch for ch in name if ch not in '\\/:*?"<>|')
                    out = os.path.join(self.pdf_base_dir, f"{safe}.pdf")
                    wb.Worksheets(name).Select(False)
                    com.export_pdf(wb, out, sheet_names=[name], one_pdf=True)
            else:
                # 파일별: 통합본 이름에서 앞부분(파일명_) 기준으로 그룹화 → 각 그룹 PDF
                groups = defaultdict(list)
                for i in range(1, wb.Worksheets.Count+1):
                    nm = wb.Worksheets(i).Name
                    key = nm.split("_")[0]
                    groups[key].append(nm)
                for key, names in groups.items():
                    wb.Worksheets(names[0]).Select(False)
                    for n in names[1:]:
                        wb.Worksheets(n).Select(True)
                    out = os.path.join(self.pdf_base_dir, f"{key}.pdf")
                    com.export_pdf(wb, out, sheet_names=None, one_pdf=True)
            wb.Close(SaveChanges=False)
            com.close()
            self.info("엑셀 + PDF 생성이 모두 완료되었습니다.")
        except Exception as e:
            self.err(f"동시 생성 중 오류: {e}")

def main():
    app = QtWidgets.QApplication(sys.argv)
    w = Main(); w.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()
